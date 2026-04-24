# ============================================================
# CONFIGURATION — Update these three paths before first use
# ============================================================

PAGES_TEMPLATE_PATH = "/path/to/your/template.pages"
PDF_OUTPUT_FOLDER   = "/Users/shrav/Desktop/tailored_resume"
MASTER_RESUME_PATH  = "master_resume.json"

# ============================================================
# IMPORTS
# ============================================================

import os
import json
import re
import subprocess
import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
from jobspy import scrape_jobs
import anthropic
from flask import Flask, request, jsonify, render_template_string
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

app    = Flask(__name__)
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

TRACKER_PATH = Path.home() / "Desktop" / "applications.xlsx"

TARGET_SEARCHES = [
  {"query": "entry level software engineer", "family": "software_engineering"},
  {"query": "new grad machine learning engineer", "family": "machine_learning"},
  {"query": "entry level data engineer", "family": "data_engineering"},
  {"query": "entry level data analyst", "family": "data_analytics"},
]

ROLE_FAMILY_PATTERNS = {
  "software_engineering": [
    r"software engineer",
    r"software developer",
    r"backend engineer",
    r"backend developer",
    r"frontend engineer",
    r"frontend developer",
    r"full stack engineer",
    r"full stack developer",
    r"fullstack engineer",
    r"fullstack developer",
    r"application engineer",
  ],
  "machine_learning": [
    r"machine learning engineer",
    r"ml engineer",
    r"ai engineer",
    r"applied ai engineer",
    r"applied scientist",
    r"computer vision engineer",
    r"nlp engineer",
  ],
  "data_engineering": [
    r"data engineer",
    r"analytics engineer",
    r"etl developer",
    r"data platform engineer",
    r"big data engineer",
  ],
  "data_analytics": [
    r"data analyst",
    r"business analyst",
    r"product analyst",
    r"business intelligence analyst",
    r"bi analyst",
    r"reporting analyst",
  ],
}

ENTRY_LEVEL_INCLUDE_PATTERNS = [
  r"entry level",
  r"entry-level",
  r"new grad",
  r"new graduate",
  r"recent grad",
  r"recent graduate",
  r"junior",
  r"associate",
  r"early career",
  r"graduate program",
  r"university grad",
  r"0\s*(?:-|to)?\s*2\+?\s*years",
  r"1\s*(?:-|to)?\s*2\+?\s*years",
  r"up to 2 years",
]

ENTRY_LEVEL_EXCLUDE_PATTERNS = [
  r"intern",
  r"internship",
  r"senior",
  r"staff",
  r"principal",
  r"lead",
  r"manager",
  r"director",
  r"vp",
  r"vice president",
  r"head of",
  r"architect",
  r"5\+?\s*years",
  r"[6-9]\+?\s*years",
  r"1[0-9]\+?\s*years",
  r"[3-9]\s*(?:-|to)\s*[9]\+?\s*years",
]

UNRELATED_ROLE_PATTERNS = [
  r"sales",
  r"account executive",
  r"marketing",
  r"designer",
  r"mechanical",
  r"electrical",
  r"civil",
  r"nurse",
  r"physician",
  r"teacher",
  r"recruiter",
  r"customer support",
  r"success manager",
  r"paralegal",
  r"technician",
]

# Cache master resume at startup
_resume_path = Path(MASTER_RESUME_PATH)
if not _resume_path.is_absolute():
    _resume_path = Path(__file__).parent / MASTER_RESUME_PATH
try:
    with open(_resume_path) as _f:
        MASTER_RESUME = json.load(_f)
except Exception as _e:
    MASTER_RESUME = None
    print(f"Warning: could not load {_resume_path}: {_e}")

# ============================================================
# HTML
# ============================================================

HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Resume Tailor v1.0</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      font-family: "Arial", sans-serif;
      font-size: 11px;
      background: #008080;
      min-height: 100vh;
      padding: 12px;
      display: flex;
    }

    .window {
      background: #c0c0c0;
      border-top: 2px solid #ffffff;
      border-left: 2px solid #ffffff;
      border-right: 2px solid #808080;
      border-bottom: 2px solid #808080;
      width: min(1400px, 100%);
      min-height: calc(100vh - 24px);
      margin: 0 auto;
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }

    .titlebar {
      background: linear-gradient(to right, #000080, #1084d0);
      padding: 3px 4px;
      display: flex;
      align-items: center;
      user-select: none;
    }
    .titlebar-left { display: flex; align-items: center; gap: 4px; }
    .titlebar-icon {
      width: 14px;
      height: 14px;
      background: #c0c0c0;
      border: 1px solid #808080;
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 9px;
      color: #000080;
      font-weight: 900;
    }
    .titlebar-text { color: #fff; font-size: 11px; font-weight: 700; }

    .menubar {
      background: #c0c0c0;
      padding: 2px 4px;
      display: flex;
      border-bottom: 1px solid #808080;
    }
    .menu-item { font-size: 11px; color: #000; padding: 2px 6px; cursor: default; }
    .menu-item:hover { background: #000080; color: #fff; }

    .window-body {
      display: flex;
      flex: 1;
      min-height: 0;
      overflow: hidden;
    }

    .left-pane {
      width: 340px;
      min-width: 340px;
      display: flex;
      flex-direction: column;
      overflow: hidden;
      background: #c0c0c0;
    }

    .pane-divider {
      width: 4px;
      background: #c0c0c0;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      flex-shrink: 0;
    }

    .right-pane {
      flex: 1;
      overflow-y: auto;
      padding: 10px;
      display: flex;
      flex-direction: column;
      gap: 10px;
      min-width: 0;
    }

    .workspace-intro {
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 10px 12px;
      background: linear-gradient(180deg, #d7d7d7 0%, #c0c0c0 100%);
    }

    .workspace-title {
      font-size: 15px;
      font-weight: 700;
      color: #000080;
      margin-bottom: 4px;
    }

    .workspace-subtitle {
      font-size: 11px;
      color: #202020;
      line-height: 1.45;
      max-width: 920px;
    }

    .workspace-grid {
      display: grid;
      grid-template-columns: minmax(0, 1.35fr) minmax(300px, 0.85fr);
      gap: 10px;
      align-items: start;
    }

    .stack-panel {
      display: flex;
      flex-direction: column;
      gap: 10px;
      min-width: 0;
    }

    .left-toolbar {
      padding: 5px 7px;
      border-bottom: 1px solid #808080;
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-shrink: 0;
    }
    .pane-title { font-size: 11px; font-weight: 700; color: #000; }

    #list-view, #detail-view {
      display: flex;
      flex-direction: column;
      flex: 1;
      overflow: hidden;
    }

    .job-list {
      flex: 1;
      overflow-y: auto;
      padding: 4px;
    }
    .no-jobs-msg {
      padding: 14px 8px;
      font-size: 11px;
      color: #808080;
      text-align: center;
      line-height: 1.5;
    }

    .job-card {
      background: #c0c0c0;
      border-top: 1px solid #fff;
      border-left: 1px solid #fff;
      border-right: 1px solid #808080;
      border-bottom: 1px solid #808080;
      padding: 6px 7px;
      margin-bottom: 4px;
      cursor: pointer;
    }
    .job-card:hover,
    .job-card.selected {
      background: #000080;
    }
    .job-card:hover .job-card-title,
    .job-card.selected .job-card-title { color: #fff; }
    .job-card:hover .job-card-company,
    .job-card.selected .job-card-company,
    .job-card:hover .job-card-location,
    .job-card.selected .job-card-location { color: #c0c0c0; }
    .job-card:hover .job-tag.time,
    .job-card.selected .job-tag.time { color: #c0c0c0; background: #000080; border-color: #1084d0; }

    .job-card-title {
      font-size: 11px;
      font-weight: 700;
      color: #000;
      margin-bottom: 1px;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    .job-card-company {
      font-size: 10px;
      color: #000;
      margin-bottom: 1px;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    .job-card-location {
      font-size: 10px;
      color: #404040;
      margin-bottom: 4px;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    .job-card-footer { display: flex; gap: 4px; flex-wrap: wrap; }

    .job-tag {
      font-size: 9px;
      padding: 1px 5px;
      border: 1px solid #808080;
    }
    .job-tag.exp  { background: #000080; color: #fff; border-color: #000080; }
    .job-tag.time { background: #c0c0c0; color: #404040; }

    .detail-toolbar {
      padding: 5px 6px;
      border-bottom: 1px solid #808080;
      display: flex;
      gap: 4px;
      flex-shrink: 0;
    }
    .detail-meta {
      padding: 6px 8px;
      border-bottom: 1px solid #808080;
      flex-shrink: 0;
    }
    .detail-job-title { font-size: 11px; font-weight: 700; color: #000; margin-bottom: 2px; }
    .detail-meta-line { font-size: 10px; color: #404040; margin-bottom: 3px; }
    .detail-jd {
      flex: 1;
      overflow-y: auto;
      padding: 8px;
      margin: 5px 6px 6px;
      font-size: 10px;
      line-height: 1.5;
      white-space: pre-wrap;
      color: #000;
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
    }

    .win-label { font-size: 11px; color: #000; margin-bottom: 2px; }
    .win-input {
      width: 100%;
      height: 22px;
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 2px 4px;
      font-size: 11px;
      font-family: Arial, sans-serif;
      color: #000;
      outline: none;
    }
    .win-textarea {
      width: 100%;
      min-height: 220px;
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 4px;
      font-size: 11px;
      font-family: Arial, sans-serif;
      color: #000;
      resize: vertical;
      outline: none;
    }
    .win-row { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }

    .win-btn {
      background: #c0c0c0;
      border-top: 2px solid #fff;
      border-left: 2px solid #fff;
      border-right: 2px solid #404040;
      border-bottom: 2px solid #404040;
      padding: 4px 10px;
      font-size: 11px;
      font-family: Arial, sans-serif;
      color: #000;
      cursor: pointer;
      min-width: 88px;
    }
    .win-btn:active {
      border-top: 2px solid #404040;
      border-left: 2px solid #404040;
      border-right: 2px solid #fff;
      border-bottom: 2px solid #fff;
    }
    .win-btn:disabled { color: #808080; cursor: default; }
    .win-btn.primary {
      font-weight: 700;
      border: 2px solid #000;
      outline: 1px solid #000;
      outline-offset: -3px;
    }
    .btn-row { display: flex; gap: 6px; justify-content: flex-start; flex-wrap: wrap; margin-top: 2px; }

    .groupbox {
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 10px 8px 8px;
      position: relative;
      min-width: 0;
    }
    .groupbox-label {
      position: absolute;
      top: -7px;
      left: 8px;
      background: #c0c0c0;
      padding: 0 3px;
      font-size: 11px;
      font-weight: 700;
      color: #000;
    }

    .result-box {
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 6px;
      font-size: 11px;
      color: #000;
      line-height: 1.6;
    }
    .score-row-win { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin: 6px 0; }
    .score-box-win {
      background: #c0c0c0;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 6px;
      text-align: center;
    }
    .score-num-win { font-size: 22px; font-weight: 700; color: #000080; }
    .score-lbl-win { font-size: 10px; color: #000; margin-top: 2px; }

    .flag-win { display: flex; align-items: flex-start; gap: 5px; font-size: 11px; color: #000; padding: 2px 0; }
    .flag-icon {
      width: 13px;
      height: 13px;
      flex-shrink: 0;
      background: #ffff00;
      border: 1px solid #808080;
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 9px;
      font-weight: 900;
      color: #000;
      margin-top: 1px;
    }

    .rec-badge {
      display: inline-block;
      background: #000080;
      color: #fff;
      font-size: 10px;
      padding: 1px 7px;
      font-weight: 700;
      text-transform: uppercase;
    }
    .rec-badge.green  { background: #008000; }
    .rec-badge.yellow { background: #808000; }
    .rec-badge.red    { background: #800000; }

    .kw-list { display: flex; flex-wrap: wrap; gap: 4px; margin-top: 3px; }
    .kw-tag  { background: #000080; color: #fff; font-size: 10px; padding: 1px 6px; }
    .kw-tag.missing { background: #800000; }
    .kw-tag.neutral { background: #808080; }

    .feedback-item { padding: 2px 0; font-size: 11px; }
    .error-box {
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 6px;
      font-size: 11px;
      color: #800000;
    }
    .pdf-path {
      font-family: "Courier New", monospace;
      font-size: 10px;
      word-break: break-all;
      color: #000080;
    }

    .win-progress {
      width: 100%;
      height: 18px;
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      overflow: hidden;
      margin: 8px 0 6px;
    }
    .win-progress-bar {
      height: 100%;
      background: repeating-linear-gradient(to right, #000080 0, #000080 10px, #1084d0 10px, #1084d0 20px);
      width: 0%;
      transition: width 0.3s;
    }

    .helper-text {
      margin-top: 6px;
      font-size: 10px;
      line-height: 1.45;
      color: #404040;
    }

    .analytics-grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 8px;
      margin-bottom: 8px;
    }

    .metric-card {
      background: #d9d9d9;
      border-top: 1px solid #fff;
      border-left: 1px solid #fff;
      border-right: 1px solid #808080;
      border-bottom: 1px solid #808080;
      padding: 8px 6px;
      min-height: 70px;
    }

    .metric-value {
      font-size: 20px;
      font-weight: 700;
      color: #000080;
      margin-bottom: 3px;
    }

    .metric-label {
      font-size: 10px;
      color: #202020;
      line-height: 1.3;
    }

    .analytics-note {
      background: #fff;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 7px;
      font-size: 10px;
      line-height: 1.5;
      color: #202020;
    }

    .results-card {
      min-height: 260px;
    }

    .results-empty {
      color: #606060;
      font-size: 11px;
      line-height: 1.5;
      padding: 6px 2px 2px;
    }

    .statusbar {
      background: #c0c0c0;
      border-top: 1px solid #808080;
      padding: 2px 6px;
      display: flex;
      gap: 6px;
    }
    .status-panel {
      flex: 1;
      background: #c0c0c0;
      border-top: 1px solid #808080;
      border-left: 1px solid #808080;
      border-right: 1px solid #fff;
      border-bottom: 1px solid #fff;
      padding: 1px 4px;
      font-size: 10px;
      color: #000;
    }

    @media (max-width: 1024px) {
      .window-body { flex-direction: column; }
      .left-pane {
        width: 100%;
        min-width: 0;
        max-height: 360px;
      }
      .pane-divider {
        width: 100%;
        height: 4px;
        border-left: 0;
        border-right: 0;
        border-top: 1px solid #808080;
        border-bottom: 1px solid #fff;
      }
      .workspace-grid { grid-template-columns: 1fr; }
    }

    @media (max-width: 640px) {
      body { padding: 0; }
      .window {
        width: 100%;
        min-height: 100vh;
      }
      .win-row,
      .analytics-grid,
      .score-row-win { grid-template-columns: 1fr; }
      .right-pane { padding: 8px; }
    }
  </style>
</head>
<body>

<div class="window">
  <div class="titlebar">
    <div class="titlebar-left">
      <div class="titlebar-icon">R</div>
      <span class="titlebar-text">Resume Tailor v1.0</span>
    </div>
  </div>

  <div class="menubar">
    <span class="menu-item">File</span>
    <span class="menu-item">Edit</span>
    <span class="menu-item">View</span>
    <span class="menu-item">Help</span>
  </div>

  <div class="window-body">
    <div class="left-pane">
      <div id="list-view">
        <div class="left-toolbar">
          <span class="pane-title">Job Listings</span>
          <button class="win-btn" id="fetch-btn" onclick="fetchJobs()">Fetch Jobs</button>
        </div>
        <div id="job-list" class="job-list">
          <div class="no-jobs-msg">Click "Fetch Jobs" to load only entry-level roles across software, ML, data engineering, and analytics.</div>
        </div>
      </div>

      <div id="detail-view" style="display:none">
        <div class="detail-toolbar">
          <button class="win-btn" onclick="showView('list')">&larr; Back</button>
          <button class="win-btn primary" onclick="useThisJob()">Use this job &rarr;</button>
        </div>
        <div class="detail-meta">
          <div id="detail-title" class="detail-job-title"></div>
          <div id="detail-company" class="detail-meta-line"></div>
          <span id="detail-exp" class="job-tag exp"></span>
        </div>
        <div id="detail-jd" class="detail-jd"></div>
      </div>
    </div>

    <div class="pane-divider"></div>

    <div class="right-pane">
      <div class="workspace-intro">
        <div class="workspace-title">Focused job search and tailoring workspace</div>
        <div class="workspace-subtitle">The left panel is filtered to entry-level roles in the four target fields. On the right, you can run a match score independently from resume generation and keep a lightweight tracker summary visible while you work.</div>
      </div>

      <div class="workspace-grid">
        <div class="groupbox">
          <span class="groupbox-label">Application brief</span>

          <div class="win-row">
            <div>
              <div class="win-label">Company: *</div>
              <input type="text" id="company" class="win-input" placeholder="e.g. Stripe" />
            </div>
            <div>
              <div class="win-label">Role: *</div>
              <input type="text" id="role" class="win-input" placeholder="e.g. Software Engineer" />
            </div>
          </div>

          <div style="margin-top:8px">
            <div class="win-label">Job URL: (optional)</div>
            <input type="text" id="url" class="win-input" placeholder="https://..." />
          </div>

          <div style="margin-top:8px">
            <div class="win-label">Job Description: *</div>
            <textarea id="jd" class="win-textarea" placeholder="Paste a job description, or click a listing on the left then 'Use this job ->'"></textarea>
          </div>

          <div class="helper-text">Use Match Score for a quick screen. Use Generate Resume when you want the tailored export and tracker update.</div>
        </div>

        <div class="stack-panel">
          <div class="groupbox">
            <span class="groupbox-label">Actions</span>
            <div class="btn-row">
              <button id="match-btn" class="win-btn primary" onclick="screenJob()">Get Match Score</button>
              <button id="generate-btn" class="win-btn primary" onclick="tailorResume()">Generate Resume</button>
              <button class="win-btn" onclick="startOver()">Clear</button>
            </div>
            <div class="win-progress"><div id="progress-bar" class="win-progress-bar"></div></div>
            <div id="action-helper" class="helper-text">No action running. Select a job or paste a description to begin.</div>
          </div>

          <div class="groupbox">
            <span class="groupbox-label">Apply analytics</span>
            <div id="analytics-content" class="results-empty">Loading tracker summary...</div>
          </div>
        </div>

        <div id="screen-card" class="groupbox results-card">
          <span class="groupbox-label">Match score</span>
          <div id="screen-content" class="results-empty">Run Match Score to evaluate fit, flags, and whether the role is worth pursuing.</div>
        </div>

        <div id="tailor-card-results" class="groupbox results-card">
          <span class="groupbox-label">Resume generation</span>
          <div id="tailor-content" class="results-empty">Generate Resume when you are ready to export a tailored PDF and log the application.</div>
        </div>
      </div>
    </div>
  </div>

  <div class="statusbar">
    <div class="status-panel" id="status-text">Ready</div>
    <div class="status-panel" style="flex:0;min-width:90px;text-align:right">localhost:5002</div>
  </div>
</div>

<script>
  let jobs = [];
  let selectedJobIdx = -1;
  let fitScore = 0;

  window.addEventListener('load', () => {
    refreshAnalytics();
  });

  function showView(view) {
    document.getElementById('list-view').style.display = view === 'list' ? 'flex' : 'none';
    document.getElementById('detail-view').style.display = view === 'detail' ? 'flex' : 'none';
  }

  async function fetchJobs() {
    const btn = document.getElementById('fetch-btn');
    btn.textContent = 'Fetching...';
    btn.disabled = true;
    setStatus('Fetching filtered entry-level listings...');
    setActionHelper('Refreshing job listings across the four target role families.');

    try {
      const res = await fetch('/jobs');
      const data = await res.json();
      if (!res.ok) throw new Error(data.error || 'Request failed');
      jobs = data.jobs || [];
      renderJobList();
      setStatus(jobs.length + ' filtered jobs found');
    } catch (err) {
      document.getElementById('job-list').innerHTML = '<div class="no-jobs-msg" style="color:#800000">&#9888; ' + escHtml(err.message) + '</div>';
      setStatus('Error fetching jobs');
    } finally {
      btn.textContent = 'Refresh';
      btn.disabled = false;
      setActionHelper('Choose a listing or paste a job description to score or generate.');
    }
  }

  function renderJobList() {
    const container = document.getElementById('job-list');
    if (!jobs.length) {
      container.innerHTML = '<div class="no-jobs-msg">No matching entry-level jobs found in the last 24 hours.</div>';
      return;
    }

    container.innerHTML = jobs.map((job, index) => `
      <div class="job-card${index === selectedJobIdx ? ' selected' : ''}" onclick="showDetail(${index})">
        <div class="job-card-title">${escHtml(job.title)}</div>
        <div class="job-card-company">${escHtml(job.company)}</div>
        <div class="job-card-location">${escHtml(job.location)}</div>
        <div class="job-card-footer">
          <span class="job-tag exp">${escHtml(job.experience)}</span>
          <span class="job-tag time">${escHtml(formatFamily(job.family))}</span>
          <span class="job-tag time">${timeAgo(job.posted)}</span>
        </div>
      </div>
    `).join('');
  }

  function showDetail(idx) {
    selectedJobIdx = idx;
    const job = jobs[idx];
    document.getElementById('detail-title').textContent = job.title;
    document.getElementById('detail-company').textContent = [job.company, job.location, formatFamily(job.family)].filter(Boolean).join(' · ');
    document.getElementById('detail-exp').textContent = job.experience;
    document.getElementById('detail-jd').textContent = job.description || 'No description was returned for this listing.';
    renderJobList();
    showView('detail');
  }

  function useThisJob() {
    if (selectedJobIdx < 0) return;
    const job = jobs[selectedJobIdx];
    document.getElementById('company').value = job.company;
    document.getElementById('role').value = job.title;
    document.getElementById('url').value = job.url;
    document.getElementById('jd').value = job.description;
    resetResults();
    setStatus('Job loaded into the form');
    setActionHelper('Selected listing copied into the form. Choose either action.');
    showView('list');
  }

  async function screenJob() {
    const company = document.getElementById('company').value.trim();
    const role = document.getElementById('role').value.trim();
    const jd = document.getElementById('jd').value.trim();

    if (!company || !role || !jd) {
      alert('Company, Role, and Job Description are required.');
      return;
    }

    const matchBtn = document.getElementById('match-btn');
    const generateBtn = document.getElementById('generate-btn');
    matchBtn.textContent = 'Scoring...';
    matchBtn.disabled = true;
    generateBtn.disabled = true;
    setStatus('Calling Claude for job screening...');
    setActionHelper('Calculating fit score and screening flags.');
    document.getElementById('screen-content').innerHTML = '<div class="results-empty">Scoring this role...</div>';

    try {
      const res = await fetch('/screen', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ jd })
      });
      const data = await res.json();
      if (!res.ok) {
        document.getElementById('screen-content').innerHTML = errorBox(data.error || 'Screening failed.');
        setStatus('Error');
        return;
      }
      fitScore = data.fit_score || 0;
      renderScreenResults(data);
      animateProgress(40);
      setStatus('Match score ready');
      setActionHelper('Score complete. Generate a resume if you want to proceed.');
    } catch (err) {
      document.getElementById('screen-content').innerHTML = errorBox('Network error: ' + err.message);
      setStatus('Error');
    } finally {
      matchBtn.textContent = 'Get Match Score';
      matchBtn.disabled = false;
      generateBtn.disabled = false;
    }
  }

  function renderScreenResults(data) {
    const recClass = {
      'strong apply': 'green',
      'apply': '',
      'borderline': 'yellow',
      'skip': 'red'
    }[data.recommendation] || '';

    const flagsHtml = (data.flags || []).map(flag =>
      `<div class="flag-win"><div class="flag-icon">!</div><span><strong>${escHtml(flag.type)}:</strong> ${escHtml(flag.detail)}</span></div>`
    ).join('');

    document.getElementById('screen-content').innerHTML = `
      <div class="result-box">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:8px;flex-wrap:wrap;margin-bottom:4px">
          <span style="font-weight:700">Fit score: ${escHtml(data.fit_score)} / 100</span>
          <span class="rec-badge ${recClass}">${escHtml(data.recommendation || 'unknown')}</span>
        </div>
        ${flagsHtml || '<span style="color:#008000">No flags detected</span>'}
      </div>`;
  }

  async function tailorResume() {
    const company = document.getElementById('company').value.trim();
    const role = document.getElementById('role').value.trim();
    const url = document.getElementById('url').value.trim();
    const jd = document.getElementById('jd').value.trim();

    if (!company || !role || !jd) {
      alert('Company, Role, and Job Description are required.');
      return;
    }

    const generateBtn = document.getElementById('generate-btn');
    const matchBtn = document.getElementById('match-btn');
    generateBtn.textContent = 'Generating...';
    generateBtn.disabled = true;
    matchBtn.disabled = true;
    animateProgress(60);
    setStatus('Tailoring resume...');
    setActionHelper('Selecting content, exporting the PDF, and updating the tracker.');
    document.getElementById('tailor-content').innerHTML = '<div class="results-empty">Generating tailored resume output...</div>';

    try {
      const res = await fetch('/tailor', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ company, role, url, jd, fit_score: fitScore })
      });
      const data = await res.json();
      if (!res.ok) {
        document.getElementById('tailor-content').innerHTML = errorBox(data.error || 'Tailoring failed.');
        setStatus('Error');
        return;
      }
      renderTailorResults(data);
      animateProgress(100);
      setStatus('Resume generation complete');
      setActionHelper('Generation complete. Tracker analytics were refreshed.');
      refreshAnalytics();
    } catch (err) {
      document.getElementById('tailor-content').innerHTML = errorBox('Network error: ' + err.message);
      setStatus('Error');
    } finally {
      generateBtn.textContent = 'Generate Resume';
      generateBtn.disabled = false;
      matchBtn.disabled = false;
    }
  }

  function renderTailorResults(data) {
    const matchedHtml = (data.ats_matched_keywords || []).map(keyword => `<span class="kw-tag">${escHtml(keyword)}</span>`).join('');
    const missingHtml = (data.ats_missing_keywords || []).map(keyword => `<span class="kw-tag missing">${escHtml(keyword)}</span>`).join('');
    const feedbackHtml = (data.recruiter_feedback || []).map(item => `<div class="feedback-item">&#8226; ${escHtml(item)}</div>`).join('');
    const projectsHtml = (data.selected_projects || []).map(project => `<span class="kw-tag">${escHtml(project)}</span>`).join('');
    const langHtml = (data.selected_languages || []).map(item => `<span class="kw-tag neutral">${escHtml(item)}</span>`).join('');
    const frameworkHtml = (data.selected_frameworks || []).map(item => `<span class="kw-tag neutral">${escHtml(item)}</span>`).join('');

    let pdfHtml = '';
    if (data.pdf_path) {
      pdfHtml = `<div style="margin-top:6px"><strong>PDF saved to:</strong><br><span class="pdf-path">${escHtml(data.pdf_path)}</span></div>`;
    } else if (data.pdf_error) {
      pdfHtml = `<div style="margin-top:6px">${errorBox('PDF export: ' + data.pdf_error)}</div>`;
    }

    let trackerHtml = '';
    if (data.tracker_updated) {
      trackerHtml = '<div style="margin-top:4px;color:#008000;font-weight:700">Logged to applications.xlsx</div>';
    } else if (data.tracker_error) {
      trackerHtml = `<div style="margin-top:4px">${errorBox('Tracker: ' + data.tracker_error)}</div>`;
    }

    document.getElementById('tailor-content').innerHTML = `
      <div class="score-row-win">
        <div class="score-box-win"><div class="score-num-win">${escHtml(data.ats_score)}</div><div class="score-lbl-win">ATS Score</div></div>
        <div class="score-box-win"><div class="score-num-win">${escHtml(data.recruiter_score)}</div><div class="score-lbl-win">Recruiter Score</div></div>
      </div>
      <div class="result-box">
        <div style="margin-bottom:4px"><strong>Projects:</strong><div class="kw-list">${projectsHtml || 'None selected'}</div></div>
        <div style="margin-bottom:4px"><strong>Languages:</strong><div class="kw-list">${langHtml || 'None selected'}</div></div>
        <div style="margin-bottom:4px"><strong>Frameworks:</strong><div class="kw-list">${frameworkHtml || 'None selected'}</div></div>
        <div style="margin-bottom:4px"><strong>Matched:</strong><div class="kw-list">${matchedHtml || 'None identified'}</div></div>
        <div style="margin-bottom:4px"><strong>Missing:</strong><div class="kw-list">${missingHtml || 'None identified'}</div></div>
        <div><strong>Feedback:</strong>${feedbackHtml || '<div class="feedback-item">No recruiter feedback returned.</div>'}</div>
        ${pdfHtml}
        ${trackerHtml}
      </div>`;
  }

  async function refreshAnalytics() {
    try {
      const res = await fetch('/analytics');
      const data = await res.json();
      if (!res.ok) throw new Error(data.error || 'Could not load analytics');
      renderAnalytics(data);
    } catch (err) {
      document.getElementById('analytics-content').innerHTML = errorBox(err.message);
    }
  }

  function renderAnalytics(data) {
    document.getElementById('analytics-content').innerHTML = `
      <div class="analytics-grid">
        <div class="metric-card"><div class="metric-value">${escHtml(data.total_applications)}</div><div class="metric-label">Total logged applications</div></div>
        <div class="metric-card"><div class="metric-value">${escHtml(data.recent_applications)}</div><div class="metric-label">Applied in the last 7 days</div></div>
        <div class="metric-card"><div class="metric-value">${escHtml(data.applied_count)}</div><div class="metric-label">Currently marked Applied</div></div>
        <div class="metric-card"><div class="metric-value">${escHtml(data.interview_count)}</div><div class="metric-label">Interview-stage roles</div></div>
        <div class="metric-card"><div class="metric-value">${escHtml(data.avg_fit_score)}</div><div class="metric-label">Average fit score</div></div>
        <div class="metric-card"><div class="metric-value">${escHtml(data.avg_ats_score)}</div><div class="metric-label">Average ATS score</div></div>
      </div>
      <div class="analytics-note">
        <strong>Average recruiter score:</strong> ${escHtml(data.avg_recruiter_score)}<br>
        <strong>Latest application:</strong> ${escHtml(data.latest_company || 'None yet')}${data.latest_role ? ' · ' + escHtml(data.latest_role) : ''}<br>
        <strong>Latest status:</strong> ${escHtml(data.latest_status || 'Not tracked yet')}
      </div>`;
  }

  function startOver() {
    document.getElementById('company').value = '';
    document.getElementById('role').value = '';
    document.getElementById('url').value = '';
    document.getElementById('jd').value = '';
    selectedJobIdx = -1;
    fitScore = 0;
    renderJobList();
    resetResults();
    animateProgress(0);
    setStatus('Ready');
    setActionHelper('No action running. Select a job or paste a description to begin.');
  }

  function resetResults() {
    document.getElementById('screen-content').innerHTML = '<div class="results-empty">Run Match Score to evaluate fit, flags, and whether the role is worth pursuing.</div>';
    document.getElementById('tailor-content').innerHTML = '<div class="results-empty">Generate Resume when you are ready to export a tailored PDF and log the application.</div>';
    document.getElementById('match-btn').textContent = 'Get Match Score';
    document.getElementById('match-btn').disabled = false;
    document.getElementById('generate-btn').textContent = 'Generate Resume';
    document.getElementById('generate-btn').disabled = false;
  }

  function setStatus(msg) { document.getElementById('status-text').textContent = msg; }
  function setActionHelper(msg) { document.getElementById('action-helper').textContent = msg; }
  function animateProgress(pct) { document.getElementById('progress-bar').style.width = pct + '%'; }

  function timeAgo(isoStr) {
    if (!isoStr) return '';
    const diff = (Date.now() - new Date(isoStr).getTime()) / 1000;
    if (diff < 60) return 'just now';
    if (diff < 3600) return Math.round(diff / 60) + 'm ago';
    if (diff < 86400) return Math.round(diff / 3600) + 'h ago';
    return Math.round(diff / 86400) + 'd ago';
  }

  function formatFamily(family) {
    return {
      software_engineering: 'Software',
      machine_learning: 'ML',
      data_engineering: 'Data Eng',
      data_analytics: 'Analytics'
    }[family] || 'Target';
  }

  function errorBox(msg) {
    return `<div class="error-box">&#9888; ${escHtml(msg)}</div>`;
  }

  function escHtml(value) {
    return String(value)
      .replace(/&/g, '&amp;')
      .replace(/</g, '&lt;')
      .replace(/>/g, '&gt;')
      .replace(/"/g, '&quot;');
  }
</script>
</body>
</html>
"""

# ============================================================
# HELPERS
# ============================================================

def esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def bullets_expr(bullets: list) -> str:
    if not bullets:
        return '""'
    return " & (return) & ".join(f'"• {esc(b)}"' for b in bullets)


def parse_claude_json(text: str) -> dict:
    text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip())
    return json.loads(text.strip())


def call_claude(prompt: str, max_tokens: int) -> dict:
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}]
    )
    return parse_claude_json(message.content[0].text)


# ============================================================
# ROUTE — Main page
# ============================================================

@app.route("/")
def index():
    return render_template_string(HTML)


# ============================================================
# ROUTE — GET /jobs
# ============================================================

def _clean(val) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val)


def _normalize_text(*parts: str) -> str:
    return " ".join(_clean(part).lower() for part in parts if _clean(part)).strip()


def _matches_any(text: str, patterns: list[str]) -> bool:
    return any(re.search(pattern, text) for pattern in patterns)


def _detect_role_family(title: str, description: str) -> str | None:
    haystack = _normalize_text(title, description)
    for family, patterns in ROLE_FAMILY_PATTERNS.items():
        if _matches_any(haystack, patterns):
            return family
    return None


def _is_entry_level_job(title: str, description: str) -> bool:
    title_text = _normalize_text(title)
    combined_text = _normalize_text(title, description)

    if _matches_any(title_text, ENTRY_LEVEL_EXCLUDE_PATTERNS):
        return False
    if _matches_any(combined_text, ENTRY_LEVEL_EXCLUDE_PATTERNS):
        return False
    if _matches_any(title_text, ENTRY_LEVEL_INCLUDE_PATTERNS):
        return True
    return _matches_any(combined_text, ENTRY_LEVEL_INCLUDE_PATTERNS)


def _keep_job(row: dict, expected_family: str) -> bool:
    title = row.get("title", "")
    description = row.get("description", "")
    company = row.get("company", "")
    title_text = _normalize_text(title)
    combined_text = _normalize_text(title, description, company)

    if not title_text:
        return False
    if _matches_any(title_text, UNRELATED_ROLE_PATTERNS):
        return False

    detected_family = _detect_role_family(title, description)
    if detected_family != expected_family:
        return False
    if _matches_any(combined_text, UNRELATED_ROLE_PATTERNS):
        return False
    return _is_entry_level_job(title, description)


def _jobspy_fetch(search: dict) -> list:
    query = search["query"]
    family = search["family"]
    try:
        df = scrape_jobs(
            site_name=["indeed", "zip_recruiter", "linkedin"],
            search_term=query,
            hours_old=24,
            results_wanted=25,
            country_indeed="USA",
            linkedin_fetch_description=True,
        )
        if df is None or df.empty:
            return []
        rows = []
        for _, row in df.iterrows():
            date_posted = row.get("date_posted")
            try:
                posted_str = date_posted.isoformat() if hasattr(date_posted, "isoformat") and not pd.isna(date_posted) else ""
            except Exception:
                posted_str = ""
            rows.append({
                "id":          _clean(row.get("id")) or _clean(row.get("job_url")),
                "title":       _clean(row.get("title")),
                "company":     _clean(row.get("company")),
                "location":    _clean(row.get("location")),
                "experience":  "Entry level",
                "posted":      posted_str,
                "url":         _clean(row.get("job_url")),
                "description": _clean(row.get("description")),
                "family":      family,
            })
        return [row for row in rows if _keep_job(row, family)]
    except Exception as e:
        print(f"jobspy error for '{query}': {e}")
        return []


@app.route("/jobs")
def get_jobs():
    with ThreadPoolExecutor(max_workers=4) as ex:
        batches = list(ex.map(_jobspy_fetch, TARGET_SEARCHES))

    seen, jobs = set(), []
    for batch in batches:
        for j in batch:
            key = j["url"] or j["id"]
            if not key or key in seen:
                continue
            seen.add(key)
            jobs.append(j)

    jobs.sort(key=lambda job: job.get("posted") or "", reverse=True)

    return jsonify({"jobs": jobs, "count": len(jobs)})


def _tracker_int(value) -> int | None:
    try:
        if value in (None, ""):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _tracker_date(value) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if not value:
        return None
    try:
        return datetime.datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _tracker_analytics() -> dict:
    summary = {
        "total_applications": 0,
        "recent_applications": 0,
        "applied_count": 0,
        "interview_count": 0,
        "avg_fit_score": 0,
        "avg_ats_score": 0,
        "avg_recruiter_score": 0,
        "latest_company": "",
        "latest_role": "",
        "latest_status": "",
    }

    if not TRACKER_PATH.exists():
        return summary

    workbook = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return summary

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    index = {header: position for position, header in enumerate(headers)}

    def cell(row, name):
        position = index.get(name)
        if position is None or position >= len(row):
            return None
        return row[position]

    today = datetime.date.today()
    fit_scores = []
    ats_scores = []
    recruiter_scores = []

    for row in rows[1:]:
        company = _clean(cell(row, "Company"))
        role = _clean(cell(row, "Position"))
        status = _clean(cell(row, "Current Status"))
        if not any([company, role, status, cell(row, "Date Applied")]):
            continue

        summary["total_applications"] += 1

        status_lower = status.lower()
        if status_lower == "applied":
            summary["applied_count"] += 1
        if "interview" in status_lower:
            summary["interview_count"] += 1

        applied_date = _tracker_date(cell(row, "Date Applied"))
        if applied_date is not None and 0 <= (today - applied_date).days <= 6:
            summary["recent_applications"] += 1

        fit_value = _tracker_int(cell(row, "Fit Score"))
        ats_value = _tracker_int(cell(row, "ATS Score"))
        recruiter_value = _tracker_int(cell(row, "Recruiter Score"))
        if fit_value is not None:
            fit_scores.append(fit_value)
        if ats_value is not None:
            ats_scores.append(ats_value)
        if recruiter_value is not None:
            recruiter_scores.append(recruiter_value)

        summary["latest_company"] = company
        summary["latest_role"] = role
        summary["latest_status"] = status or "Unknown"

    if fit_scores:
        summary["avg_fit_score"] = round(sum(fit_scores) / len(fit_scores))
    if ats_scores:
        summary["avg_ats_score"] = round(sum(ats_scores) / len(ats_scores))
    if recruiter_scores:
        summary["avg_recruiter_score"] = round(sum(recruiter_scores) / len(recruiter_scores))

    return summary


@app.route("/analytics")
def analytics():
    try:
        return jsonify(_tracker_analytics())
    except Exception as e:
        return jsonify({"error": f"Analytics failed: {e}"}), 500


# ============================================================
# ROUTE — POST /screen
# ============================================================

@app.route("/screen", methods=["POST"])
def screen():
    body = request.get_json(silent=True) or {}
    jd   = (body.get("jd") or "").strip()
    if not jd:
        return jsonify({"error": "Job description is required."}), 400

    prompt = f"""You are a job application screener. Analyze this job description and return a JSON object with EXACTLY these fields:

- fit_score: integer 0-100. Holistic estimate of how good a fit this role is for a recent CS graduate with ~2 years of project experience.
- flags: array of objects, each with "type" and "detail". Detect these types only:
    * "no_sponsorship" — JD says no visa sponsorship or requires authorization to work in the US
    * "seniority" — role clearly requires 5+ years of experience, or is Staff / Principal / Director / VP level
    * "security_clearance" — role requires any security clearance
  If none apply, return an empty array.
- flag_count: integer equal to the length of the flags array.
- recommendation: exactly one of: "strong apply", "apply", "borderline", "skip"
  Use this scale — strong apply (fit_score 80+, no bad flags), apply (60-79), borderline (40-59 or has one flag), skip (below 40 or has 2+ flags or security clearance).

Return ONLY the raw JSON object. No markdown, no explanation, no code fences.

Job Description:
{jd}"""

    try:
        return jsonify(call_claude(prompt, max_tokens=512))
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Claude returned invalid JSON: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Screening failed: {e}"}), 500


# ============================================================
# ROUTE — POST /tailor
# ============================================================

@app.route("/tailor", methods=["POST"])
def tailor():
    body      = request.get_json(silent=True) or {}
    company   = (body.get("company")   or "").strip()
    role      = (body.get("role")      or "").strip()
    url       = (body.get("url")       or "").strip()
    jd        = (body.get("jd")        or "").strip()
    fit_score = body.get("fit_score", 0)

    if not company or not role or not jd:
        return jsonify({"error": "Company, role, and job description are required."}), 400
    if MASTER_RESUME is None:
        return jsonify({"error": "master_resume.json could not be loaded. Check the path in the config block."}), 500

    prompt = f"""You are a resume tailoring expert. Given a job description and a master resume, select the best content for a targeted one-page resume.

IMPORTANT: The final resume must fit on ONE PAGE. Err on the side of selecting fewer items. It is better to have breathing room than to cram content.

The resume has two separate skills lines:
  Line 1 — "Programming languages: [values]"
  Line 2 — "Frameworks/Technologies: [values]"
Select separately from the languages list and the frameworks list in the master resume.

Return a JSON object with EXACTLY these fields:
- selected_languages: ordered array of language names from skills.languages. Pick the most relevant, up to 5.
- selected_frameworks: ordered array of framework/tool names from skills.frameworks. Pick the most relevant, up to 10.
- selected_projects: array of EXACTLY 3 project names from the master resume. Pick the 3 best matches.
- ats_score: integer 0-100. How well the selected content matches ATS keywords in the JD.
- ats_missing_keywords: array of important JD keywords NOT found in the selected content. List up to 8.
- ats_matched_keywords: array of JD keywords that DO appear in the selected content. List up to 12.
- recruiter_score: integer 0-100. How compelling this tailored resume would appear to a human recruiter.
- recruiter_feedback: array of exactly 2–3 short, specific, actionable feedback strings.

Return ONLY the raw JSON object. No markdown, no explanation, no code fences.

Job Description:
{jd}

Master Resume (JSON):
{json.dumps(MASTER_RESUME, indent=2)}"""

    try:
        result = call_claude(prompt, max_tokens=1024)
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Claude returned invalid JSON: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Tailoring API call failed: {e}"}), 500

    project_map            = {p["name"]: p for p in MASTER_RESUME.get("projects", [])}
    selected_project_names = result.get("selected_projects", [])[:3]
    selected_projects      = [project_map[n] for n in selected_project_names if n in project_map]
    while len(selected_projects) < 3:
        selected_projects.append({"name": "Project", "bullets": []})

    languages_str  = ", ".join(result.get("selected_languages", []))
    frameworks_str = ", ".join(result.get("selected_frameworks", []))

    date_label   = datetime.date.today().strftime("%B%-d")
    company_safe = re.sub(r"[^A-Za-z0-9]", "", company)
    job_folder   = Path(PDF_OUTPUT_FOLDER) / f"{company_safe}_{date_label}"
    job_folder.mkdir(parents=True, exist_ok=True)
    pdf_path   = str(job_folder / "ShravanthiMurugesan_Resume.pdf")
    pages_path = str(job_folder / "ShravanthiMurugesan_Resume.pages")

    pdf_error = None
    try:
        run_pages_export(languages_str, frameworks_str, selected_projects, pdf_path, pages_path)
    except Exception as e:
        pdf_error = str(e)
        pdf_path  = None

    tracker_error = None
    try:
        log_to_tracker(company, role, url, jd, fit_score,
                       result.get("ats_score", 0), result.get("recruiter_score", 0),
                       pdf_path or "")
    except Exception as e:
        tracker_error = str(e)

    return jsonify({
        **result,
        "pdf_path":        pdf_path,
        "pdf_error":       pdf_error,
        "tracker_updated": tracker_error is None,
        "tracker_error":   tracker_error,
    })


# ============================================================
# HELPER — AppleScript: fill Pages template, export PDF
# Template tokens: {{skills_languages}}, {{skills_frameworks}},
#   {{project1_name}}, {{project1_bullets}}, (same for 2 and 3)
# ============================================================

def run_pages_export(languages_str, frameworks_str, projects, pdf_path, pages_path):
    p1, p2, p3 = projects[0], projects[1], projects[2]

    script = f"""
tell application "Pages"
    open POSIX file "{esc(PAGES_TEMPLATE_PATH)}"
    delay 2
    tell document 1
        replace "{{{{skills_languages}}}}"  with "{esc(languages_str)}"  replacing all true
        replace "{{{{skills_frameworks}}}}" with "{esc(frameworks_str)}" replacing all true
        replace "{{{{project1_name}}}}"     with "{esc(p1['name'])}"     replacing all true
        replace "{{{{project2_name}}}}"     with "{esc(p2['name'])}"     replacing all true
        replace "{{{{project3_name}}}}"     with "{esc(p3['name'])}"     replacing all true
        set p1b to {bullets_expr(p1.get('bullets', []))}
        set p2b to {bullets_expr(p2.get('bullets', []))}
        set p3b to {bullets_expr(p3.get('bullets', []))}
        replace "{{{{project1_bullets}}}}" with p1b replacing all true
        replace "{{{{project2_bullets}}}}" with p2b replacing all true
        replace "{{{{project3_bullets}}}}" with p3b replacing all true
        save in POSIX file "{esc(pages_path)}"
        export to POSIX file "{esc(pdf_path)}" as PDF
    end tell
end tell
"""

    result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "AppleScript failed with no error message.")


# ============================================================
# HELPER — Log one row to applications.xlsx
# ============================================================

def log_to_tracker(company, role, url, jd, fit_score, ats_score, recruiter_score, pdf_path):
    columns    = ["Company", "Position", "Referral?", "Date Applied", "Current Status",
                  "URL", "JD", "Fit Score", "ATS Score", "Recruiter Score", "PDF Path"]
    col_widths = [22, 32, 12, 14, 16, 40, 60, 10, 10, 16, 50]

    try:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        for col_idx, header in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.append([company, role, "", datetime.date.today().strftime("%Y-%m-%d"),
               "Applied", url, jd, fit_score, ats_score, recruiter_score, pdf_path])
    ws.cell(row=ws.max_row, column=7).alignment = Alignment(wrap_text=True)
    wb.save(TRACKER_PATH)


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    print("Resume Tailor running at http://localhost:5002")
    print("Press Ctrl+C to stop.")
    app.run(debug=False, port=5002)
